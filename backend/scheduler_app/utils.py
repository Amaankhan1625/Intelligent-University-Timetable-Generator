from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import datetime

def export_timetable_pdf(timetable):
    """Export timetable to PDF format (Monday-Friday only)."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=36, leftMargin=36,
                           topMargin=36, bottomMargin=18)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        fontSize=14,
        alignment=1,
        spaceAfter=8
    )

    story = []
    title = Paragraph(f"{timetable.name}", title_style)
    story.append(title)
    story.append(Spacer(1, 6))

    # Get classes
    classes = timetable.classes.all().select_related('course', 'instructor', 'room', 'meeting_time', 'section')

    # Days: Monday-Friday only
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

    # Collect unique time slots (start-end string)
    time_slots = []
    for cls in classes:
        if not cls.meeting_time:
            continue
        start = cls.meeting_time.start_time
        end = cls.meeting_time.end_time
        slot = f"{start.strftime('%H:%M')}-{end.strftime('%H:%M')}"
        if slot not in time_slots:
            time_slots.append(slot)

    time_slots.sort()

    # Build header
    table_data = []
    header_row = ['Time'] + days
    table_data.append(header_row)

    for time_slot in time_slots:
        row = [time_slot]
        start_time_obj = datetime.datetime.strptime(time_slot.split('-')[0], '%H:%M').time()
        for day in days:
            cell_content_lines = []
            # Filter classes matching day and start_time
            day_classes = classes.filter(meeting_time__day=day, meeting_time__start_time=start_time_obj)
            for cls in day_classes:
                course_id = getattr(cls.course, 'course_id', '')
                instructor = getattr(cls.instructor, 'name', '')
                room = getattr(cls.room, 'room_number', '')
                section = getattr(cls.section, 'section_id', '')
                cell_content_lines.append(f"{course_id} / {cls.course.course_name}")
                cell_content_lines.append(f"{instructor} | {room} | {section}")
                # blank line between entries
                cell_content_lines.append("")
            cell_content = "\n".join(cell_content_lines).strip()
            row.append(cell_content)
        table_data.append(row)

    if len(table_data) == 1:
        # No slots — show a friendly message
        story.append(Paragraph("No scheduled classes to export.", styles['Normal']))
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f4f4f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))

    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def export_timetable_excel(timetable):
    """Export timetable to Excel (Monday-Friday)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    classes = timetable.classes.all().select_related('course', 'instructor', 'room', 'meeting_time', 'section')

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    time_slots = []
    for cls in classes:
        if not cls.meeting_time:
            continue
        slot = f"{cls.meeting_time.start_time.strftime('%H:%M')}-{cls.meeting_time.end_time.strftime('%H:%M')}"
        if slot not in time_slots:
            time_slots.append(slot)
    time_slots.sort()

    ws['A1'] = timetable.name
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center_alignment

    headers = ['Time'] + days
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    if not time_slots:
        ws.cell(row=5, column=1, value="No scheduled classes")
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    for row_idx, time_slot in enumerate(time_slots, 4):
        ws.cell(row=row_idx, column=1, value=time_slot).alignment = center_alignment
        start_time_obj = datetime.datetime.strptime(time_slot.split('-')[0], '%H:%M').time()

        for col_idx, day in enumerate(days, 2):
            cell_lines = []
            day_classes = classes.filter(meeting_time__day=day, meeting_time__start_time=start_time_obj)
            for cls in day_classes:
                cid = getattr(cls.course, 'course_id', '')
                cname = getattr(cls.course, 'course_name', '')
                instr = getattr(cls.instructor, 'name', '')
                room = getattr(cls.room, 'room_number', '')
                section = getattr(cls.section, 'section_id', '')
                cell_lines.append(f"{cid} {cname}")
                cell_lines.append(f"{instr} | {room} | {section}")
                cell_lines.append("")
            cell_value = "\n".join(cell_lines).strip()
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = center_alignment

    # Auto column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
